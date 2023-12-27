from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep

from selenium.webdriver.support.wait import WebDriverWait # ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec # beklenen koşulları webdriver a verilen
from selenium.webdriver.common.action_chains import ActionChains # zincir gibi aksiyonları sıraya koymamıza yardımcı oluyor

import pytest
from constants import globalConstants as gc
import openpyxl

class TestSauceScenariov3:

    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(gc.BASE_URL)
        self.driver.maximize_window()
    
    def teardown_method(self):
        self.driver.quit()
    
    def get_Login_Data(login_type):
        excel = openpyxl.load_workbook(gc.LOGINDATA_xlsx)
        match login_type:
            case 0:
                sheet =excel["emptyPassLogin"]

                rows = sheet.max_row
                data =[]
                for i in range(2, rows+1): 
                    username = sheet.cell(i,1).value 
                    data.append((username))
                return data
            case 1:
                return [("locked_out_user","secret_sauce")]
            case 2:
                sheet =excel["notRegisteredLogin"]

                rows = sheet.max_row
                data =[]
                for i in range(2, rows+1): 
                    username = sheet.cell(i,1).value 
                    password = sheet.cell(i,2).value
                    data.append((username,password)) 
                return data
            case 3:
                return [("standard_user","secret_sauce")] 
               
    # LOGIN, TEST CASE 1    
    # check expected result of the empty username&password login 
    @pytest.mark.skip
    #@pytest.mark.parametrize("username,password", get_Login_Data(0))
    def test_empty_username_pass_login(self):

        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.LOGIN_BUTTON_ID)))
        loginButton.click()

        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gc.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Username is required"

    # LOGIN, TEST CASE 2        
    # check expected result of the empty password login
    @pytest.mark.skip 
    @pytest.mark.parametrize("username",get_Login_Data(0)) # login_type=1
    def test_empty_pass_login(self,username):

        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.LOGIN_BUTTON_ID)))
        loginButton.click()

        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gc.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Password is required"
    
    # LOGIN, TEST CASE 3    
    # check expected result of the locked out user login 
    @pytest.mark.skip
    @pytest.mark.parametrize("username,password",get_Login_Data(1))# login_type 1
    def test_locked_login(self,username,password): 

        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(password)

        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.LOGIN_BUTTON_ID)))
        loginButton.click()

        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gc.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Sorry, this user has been locked out."
    
    # LOGIN, TEST CASE 4    
    # not registered user login
    @pytest.mark.skip
    @pytest.mark.parametrize("username,password",get_Login_Data(2))# login_type 2
    def test_not_registered_user_login(self, username, password):

        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(password)

        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.LOGIN_BUTTON_ID)))
        loginButton.click()

        errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gc.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
    
    # LOGIN, TEST CASE 5    
    # succesfull login
    # check expected results of the standard user login 
    @pytest.mark.skip
    @pytest.mark.parametrize("username,password",get_Login_Data(3))# login_type 3
    def test_standard_user_login(self,username,password): #login_type 3
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(password)

        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.LOGIN_BUTTON_ID)))
        loginButton.click()

        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"inventory_container")))

        current_page_test_result= self.driver.current_url 
        assert current_page_test_result == gc.BASE_URL+"/inventory.html", "not expected page url"
    
        listOfCourses = self.driver.find_elements(By.CLASS_NAME, "inventory_item")
        assert len(listOfCourses) == 6,  "number of courses is different from 6"
        
    ## Test cases for product List Page ##
    # product List Page, TEST CASE 1    
    # add all the products in the page to the shopping cart and check if they added to the shopping cart succesfully
    @pytest.mark.skip
    def test_addToCart(self):
        self.test_standard_user_login("standard_user","secret_sauce")

        # find all addToCart buttons in the product lişst page
        addToCartButtonList = WebDriverWait(self.driver,5).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "btn")))
        actionChains = ActionChains(self.driver)
        addToCart_itemNameLists =[]

        # find all inventory names in the product list page
        itemNameList = WebDriverWait(self.driver,5).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "inventory_item_name")))
        i=0

        # move for each addToCart button in the page
        for addToCart in addToCartButtonList:
            # move to current add to cart button and click 
            actionChains.move_to_element(addToCart) # sayfayı elementin olduğu yere taşı
            actionChains.click()
            actionChains.perform()
            
            # add clicked button item name into a list
            addToCart_itemNameLists.append(itemNameList[i].text)

            # create remove button name with item name, e.g. remove-sauce-labs-backpack
            remove_btn_name ="remove-"+str.lower(itemNameList[i].text).replace(" ","-")
            # find remove button by name and check its text if equal Remove
            removeButton =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, remove_btn_name)))
            assert removeButton.text == "Remove" ,{"text of Button ",remove_btn_name," not equal 'Remove'"}
            i+=1
            # class shopping_cart_badge should be visible and available here

        # find shopping cart link in the page and click
        shopping_cart_link =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CLASS_NAME, "shopping_cart_link")))
        shopping_cart_link.click()
       
        # find shopping cart page title and check if it equals "Your Cart"
        shopping_cart_page_title =  WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CLASS_NAME, "title")))
        assert shopping_cart_page_title.text == "Your Cart", "title of the page is not equal 'Your Cart'"
        
        # find item names that are added to the shopping Cart 
        shoppingCart_itemNameList = WebDriverWait(self.driver,5).until(ec.presence_of_all_elements_located((By.CLASS_NAME, "inventory_item_name")))
        i=0
        for shoppingCart_itemName in shoppingCart_itemNameList:
            assert shoppingCart_itemName.text == addToCart_itemNameLists[i] ,{"item ",shoppingCart_itemName.text," could not find in the shopping cart"}
            i+=1
    
    # product List Page, TEST CASE 2 
    # check go to product details page by clicking item name
    @pytest.mark.skip
    def test_goToproductDetailsByItemName(self):
        self.test_standard_user_login("standard_user","secret_sauce")
        # find first item name and get its text
        itemName = WebDriverWait(self.driver,5).until(ec.presence_of_element_located((By.CLASS_NAME, "inventory_item_name")))
        itemNameText = itemName.text
        
        # find item name text link and click
        itemLink= WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, "item_4_title_link")))
        itemLink.click()
        
        # check if text of item name link and item name that is shown in the product detail page match 
        itemDetailsName = WebDriverWait(self.driver,5).until(ec.presence_of_element_located((By.CLASS_NAME, "inventory_details_name")))
        assert itemDetailsName.text == itemNameText, "Item Names are not equal"
    
    # product List Page, TEST CASE 3 
    # check go to product details page by clicking item image name
    #@pytest.mark.skip
    def test_goToproductDetailsByItemImage(self):
        self.test_standard_user_login("standard_user","secret_sauce")
        # find first item name of the item image and get its text
        itemImage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='item_4_img_link']/img")))
        itemImageName = itemImage.accessible_name
        
        # find first item image link and click
        itemImageLink= WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, "item_4_img_link")))
        itemImageLink.click()
        
        # check if text of item image link and item image name that is shown in the product detail page match 
        itemDetailsImage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CLASS_NAME, "inventory_details_img")))
        assert itemDetailsImage.accessible_name == itemImageName, "Item Images does not match"